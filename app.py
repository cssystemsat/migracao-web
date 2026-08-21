import base64
import hashlib
import io
import json
import os
import threading
import time
import uuid

import firebase_admin
import openpyxl
import pandas as pd
import requests
from firebase_admin import credentials, firestore
from google.cloud.firestore_v1.base_query import FieldFilter
from flask import Flask, Response, jsonify, redirect, render_template, request, session, url_for
from werkzeug.security import check_password_hash, generate_password_hash

import conversorkml

app = Flask(__name__)
app.secret_key = os.environ.get("MIGRACAO_SECRET_KEY", os.urandom(24))
app.config["SEND_FILE_MAX_AGE_DEFAULT"] = 0

# Sobe 0.1 a cada edição publicada (2.0 -> 2.1 -> 2.2 ...); só sobe o inteiro quando pedido.
APP_VERSION = "2.15"

BASE_URL = "https://integration.systemsatx.com.br"

# --- FIREBASE / FIRESTORE (banco de dados) ---
# Credencial vem inteira (o JSON da chave de serviço) numa única variável de
# ambiente, tanto local quanto no Render, pra não depender de upload de arquivo.
_FIREBASE_CRED_JSON = os.environ.get("FIREBASE_CREDENTIALS_JSON")
if not _FIREBASE_CRED_JSON:
    raise RuntimeError(
        "Variável de ambiente FIREBASE_CREDENTIALS_JSON não definida. "
        "Configure-a com o conteúdo do arquivo de chave de serviço do Firebase."
    )
_firebase_cred = credentials.Certificate(json.loads(_FIREBASE_CRED_JSON))
firebase_admin.initialize_app(_firebase_cred)
db = firestore.client()
CREDENCIAIS_COLLECTION = "credenciais"

# --- CONTAS DE ACESSO À PRÓPRIA FERRAMENTA (protege o app quando publicado na internet) ---
# Guardadas no Firestore (coleção "app_usuarios"); cada uma tem usuário, hash de senha
# (nunca a senha em texto puro) e uma flag admin (só admins gerenciam outros usuários).
# Sem nenhum usuário cadastrado ainda, a tela de login vira um assistente de primeiro
# acesso que cria o usuário administrador inicial.
APP_USUARIOS_COLLECTION = "app_usuarios"


def _sem_usuarios_cadastrados():
    return len(list(db.collection(APP_USUARIOS_COLLECTION).limit(1).stream())) == 0


def _buscar_usuario_app(usuario):
    usuario_norm = usuario.strip().lower()
    docs = db.collection(APP_USUARIOS_COLLECTION).where(
        filter=FieldFilter("usuario_norm", "==", usuario_norm)
    ).limit(1).stream()
    for d in docs:
        return dict(d.to_dict(), id=d.id)
    return None


def _criar_usuario_app(usuario, senha, admin=False):
    doc_ref = db.collection(APP_USUARIOS_COLLECTION).document()
    dados = {
        "usuario": usuario,
        "usuario_norm": usuario.strip().lower(),
        "senha_hash": generate_password_hash(senha),
        "admin": admin,
    }
    doc_ref.set(dados)
    return dict(dados, id=doc_ref.id)


def _sessao_login_app(usuario_doc):
    session["app_ok"] = True
    session["app_usuario_id"] = usuario_doc["id"]
    session["app_usuario_nome"] = usuario_doc["usuario"]
    session["app_usuario_admin"] = bool(usuario_doc.get("admin"))


def _sou_admin():
    return session.get("app_usuario_admin") is True


@app.before_request
def exigir_login_app():
    if request.endpoint in ("login_app", "logout_app", "static"):
        return
    if not session.get("app_ok"):
        return redirect(url_for("login_app"))


@app.route("/login-app", methods=["GET", "POST"])
def login_app():
    modo_setup = _sem_usuarios_cadastrados()
    erro = None
    if request.method == "POST":
        usuario = request.form.get("usuario", "").strip()
        senha = request.form.get("senha", "")

        if modo_setup:
            confirmar = request.form.get("confirmar_senha", "")
            if not usuario or not senha:
                erro = "Informe usuário e senha."
            elif senha != confirmar:
                erro = "As senhas não conferem."
            elif len(senha) < 4:
                erro = "A senha deve ter pelo menos 4 caracteres."
            else:
                novo = _criar_usuario_app(usuario, senha, admin=True)
                _sessao_login_app(novo)
                return redirect(url_for("index"))
        else:
            doc = _buscar_usuario_app(usuario) if usuario else None
            if doc and check_password_hash(doc["senha_hash"], senha):
                _sessao_login_app(doc)
                return redirect(url_for("index"))
            erro = "Usuário ou senha incorretos."

    return render_template("login_app.html", erro=erro, app_version=APP_VERSION, modo_setup=modo_setup)


@app.route("/logout-app")
def logout_app():
    session.pop("app_ok", None)
    session.pop("app_usuario_id", None)
    session.pop("app_usuario_nome", None)
    session.pop("app_usuario_admin", None)
    return redirect(url_for("login_app"))


@app.route("/api/app-usuarios", methods=["GET"])
def listar_app_usuarios():
    if not _sou_admin():
        return jsonify(ok=False, error="Apenas administradores podem gerenciar usuários."), 403
    docs = db.collection(APP_USUARIOS_COLLECTION).stream()
    lista = [{"id": d.id, "usuario": d.to_dict().get("usuario"), "admin": bool(d.to_dict().get("admin"))} for d in docs]
    return jsonify(ok=True, usuarios=lista)


@app.route("/api/app-usuarios", methods=["POST"])
def criar_app_usuario():
    if not _sou_admin():
        return jsonify(ok=False, error="Apenas administradores podem gerenciar usuários."), 403
    data = request.get_json(force=True) or {}
    usuario = str(data.get("usuario", "")).strip()
    senha = str(data.get("senha", ""))
    admin = bool(data.get("admin"))
    if not usuario or not senha:
        return jsonify(ok=False, error="Informe usuário e senha."), 400
    if len(senha) < 4:
        return jsonify(ok=False, error="A senha deve ter pelo menos 4 caracteres."), 400
    if _buscar_usuario_app(usuario):
        return jsonify(ok=False, error="Já existe um usuário com esse nome."), 400
    novo = _criar_usuario_app(usuario, senha, admin)
    return jsonify(ok=True, usuario=dict(id=novo["id"], usuario=novo["usuario"], admin=novo["admin"]))


@app.route("/api/app-usuarios/<usuario_id>", methods=["PUT"])
def editar_app_usuario(usuario_id):
    if not _sou_admin():
        return jsonify(ok=False, error="Apenas administradores podem gerenciar usuários."), 403
    data = request.get_json(force=True) or {}
    doc_ref = db.collection(APP_USUARIOS_COLLECTION).document(usuario_id)
    doc = doc_ref.get()
    if not doc.exists:
        return jsonify(ok=False, error="Usuário não encontrado."), 404

    atual = doc.to_dict()
    novo_nome = str(data.get("usuario", atual["usuario"])).strip()
    nova_senha = str(data.get("senha", "")).strip()
    novo_admin = bool(data.get("admin", atual.get("admin")))

    if not novo_nome:
        return jsonify(ok=False, error="Informe o nome do usuário."), 400

    outro = _buscar_usuario_app(novo_nome)
    if outro and outro["id"] != usuario_id:
        return jsonify(ok=False, error="Já existe um usuário com esse nome."), 400

    # Não deixa remover o último administrador (senão ninguém mais gerencia usuários).
    if atual.get("admin") and not novo_admin:
        outros_admins = db.collection(APP_USUARIOS_COLLECTION).where(filter=FieldFilter("admin", "==", True)).stream()
        if sum(1 for a in outros_admins if a.id != usuario_id) == 0:
            return jsonify(ok=False, error="Precisa existir pelo menos um administrador."), 400

    dados = {
        "usuario": novo_nome,
        "usuario_norm": novo_nome.lower(),
        "admin": novo_admin,
        "senha_hash": atual["senha_hash"],
    }
    if nova_senha:
        if len(nova_senha) < 4:
            return jsonify(ok=False, error="A senha deve ter pelo menos 4 caracteres."), 400
        dados["senha_hash"] = generate_password_hash(nova_senha)

    doc_ref.set(dados)

    if usuario_id == session.get("app_usuario_id"):
        session["app_usuario_nome"] = novo_nome
        session["app_usuario_admin"] = novo_admin

    return jsonify(ok=True, usuario=dict(id=usuario_id, usuario=novo_nome, admin=novo_admin))


@app.route("/api/app-usuarios/<usuario_id>", methods=["DELETE"])
def excluir_app_usuario(usuario_id):
    if not _sou_admin():
        return jsonify(ok=False, error="Apenas administradores podem gerenciar usuários."), 403
    if usuario_id == session.get("app_usuario_id"):
        return jsonify(ok=False, error="Você não pode excluir o próprio usuário enquanto está logado."), 400
    doc_ref = db.collection(APP_USUARIOS_COLLECTION).document(usuario_id)
    doc = doc_ref.get()
    if not doc.exists:
        return jsonify(ok=False, error="Usuário não encontrado."), 404
    if doc.to_dict().get("admin"):
        outros_admins = db.collection(APP_USUARIOS_COLLECTION).where(filter=FieldFilter("admin", "==", True)).stream()
        if sum(1 for a in outros_admins if a.id != usuario_id) == 0:
            return jsonify(ok=False, error="Precisa existir pelo menos um administrador."), 400
    doc_ref.delete()
    return jsonify(ok=True)


@app.route("/api/app-usuario/senha", methods=["POST"])
def trocar_minha_senha_app():
    usuario_id = session.get("app_usuario_id")
    if not usuario_id:
        return jsonify(ok=False, error="Não autenticado."), 401
    data = request.get_json(force=True) or {}
    senha_atual = str(data.get("senha_atual", ""))
    senha_nova = str(data.get("senha_nova", ""))
    if len(senha_nova) < 4:
        return jsonify(ok=False, error="A nova senha deve ter pelo menos 4 caracteres."), 400
    doc_ref = db.collection(APP_USUARIOS_COLLECTION).document(usuario_id)
    doc = doc_ref.get()
    if not doc.exists:
        return jsonify(ok=False, error="Usuário não encontrado."), 404
    if not check_password_hash(doc.to_dict()["senha_hash"], senha_atual):
        return jsonify(ok=False, error="Senha atual incorreta."), 400
    doc_ref.update({"senha_hash": generate_password_hash(senha_nova)})
    return jsonify(ok=True)

# --- DEFINIÇÃO DE TIPOS CONFORME DOCUMENTAÇÃO SSX ---
# Comparado pelo caminho COMPLETO do parâmetro (com dot notation), não só pelo nome
# final, para não confundir campos com o mesmo nome em contextos diferentes
# (ex.: "PhoneNumber" do Cliente é texto, mas "Tracker1.Simcard1.PhoneNumber" é int32).
INT_FIELDS = [
    "ModelYear", "FabricationYear", "Fuel", "IdMapIcon", "IdMapIconColor",
    "IgnitionStatus", "OperationalStatus", "GPSStatus", "WarningStatus",
    "IdModelTracker", "TypeOrganizationalUnit", "Language", "Country", "TimeZone",
    "Tracker1.Simcard1.CountryCode", "Tracker1.Simcard1.AreaCode", "Tracker1.Simcard1.PhoneNumber",
    "Tracker1.Simcard2.CountryCode", "Tracker1.Simcard2.AreaCode", "Tracker1.Simcard2.PhoneNumber",
    "Tracker2.Simcard1.CountryCode", "Tracker2.Simcard1.AreaCode", "Tracker2.Simcard1.PhoneNumber",
    "Tracker2.Simcard2.CountryCode", "Tracker2.Simcard2.AreaCode", "Tracker2.Simcard2.PhoneNumber",
]
BOOL_FIELDS = ["Active", "ChangePasswordNextLogin", "SendPasswordEmail"]

PARAMS = {
    "cliente": {
        "titulo": "Cliente",
        "endpoint": "/Administration/Client/Insert",
        "defaults": {"Language": 1, "Country": 29, "TimeZone": 31},
        "campos": [
            "ClientIntegrationCode", "ClientTemplateIntegrationCode",
            "OrganizationalUnitIntegrationCode", "Code", "ClientType",
            "TradingName", "CompanyName", "DocumentNumber", "RegisterNumber",
            "CustomerSupportProcedure", "UserName", "Login", "Password",
            "UserProfileTemplateIntegrationCode", "PhoneNumber", "CellPhoneNumber",
        ],
    },
    "uo": {
        "titulo": "UO",
        "endpoint": "/Administration/OrganizationalUnit/Insert",
        "defaults": {},
        "campos": [
            "Name", "OrganizationalUnitIntegrationCode", "TypeOrganizationalUnit",
            "ParentOrganizationalUnitIntegrationCode", "ClientIntegrationCode",
            "Note", "Active", "ExternalIntegrationCode",
        ],
    },
    "usuario": {
        "titulo": "Usuário",
        "endpoint": "/Administration/User/Insert",
        "defaults": {},
        "campos": [
            "Name", "Login", "Password", "Email", "ClientIntegrationCode",
            "OrganizationalUnitIntegrationCode", "ProfileTemplateIntegrationCode",
            "UserIntegrationCode", "Active", "Note", "ExternalIntegrationCode",
            "ChangePasswordNextLogin", "SendPasswordEmail", "PhoneNumber",
            "DocumentNumber", "Language", "TimeZone", "UserType",
        ],
    },
    "veiculo": {
        "titulo": "Veículo",
        "endpoint": "/Administration/Vehicle/Insert",
        "defaults": {"IgnitionStatus": 1, "OperationalStatus": 1, "GPSStatus": 1, "WarningStatus": 1},
        "campos": [
            "VehicleIntegrationCode", "ClientIntegrationCode", "Identification",
            "LicensePlate", "ChassiNumber", "RenavamNumber", "Color",
            "FederalState", "City", "ModelYear", "FabricationYear", "Fuel",
            "QRCode", "FipeCode", "IdMapIcon", "IdMapIconColor",
            "ClientVehicleIntegrationCode",
            "Tracker1.TrackerIntegrationCode", "Tracker1.IdTracker",
            "Tracker1.TrackerTemplateIntegrationCode", "Tracker1.TrackerIMEI",
            "Tracker1.Simcard1.ICCID", "Tracker1.Simcard1.APN",
            "Tracker1.Simcard1.CountryCode", "Tracker1.Simcard1.AreaCode",
            "Tracker1.Simcard1.PhoneNumber",
        ],
    },
}

# Rótulos amigáveis exibidos para o usuário no mapeamento de colunas.
# A chamada da API sempre usa o nome interno (chave do PARAMS["campos"]);
# aqui só trocamos o texto mostrado na tela. Campo sem entrada aqui usa o nome interno.
LABELS = {
    "cliente": {
        "ClientIntegrationCode": "Código de integração do cliente",
        "ClientTemplateIntegrationCode": "Template do cliente",
        "OrganizationalUnitIntegrationCode": "Código da Unidade organizacional",
        "Code": "Código do cliente",
        "ClientType": "Tipo do cliente",
        "TradingName": "Nome do cliente",
        "CompanyName": "Razão Social",
        "DocumentNumber": "CPF/CNPJ",
        "RegisterNumber": "IE / RG",
        "CustomerSupportProcedure": "Procedimento de atendimento",
        "UserName": "Nome do Usuário master",
        "Login": "Login do Cliente",
        "Password": "Senha do Cliente",
        "UserProfileTemplateIntegrationCode": "Template de perfil de acesso",
        "PhoneNumber": "Telefone do Cliente",
        "CellPhoneNumber": "Celular do Cliente",
    },
    "veiculo": {
        "VehicleIntegrationCode": "Código de integração do veículo",
        "ClientIntegrationCode": "Código de integração do cliente",
        "Identification": "Identificação do Veículo",
        "LicensePlate": "Placa",
        "ChassiNumber": "Chassi",
        "RenavamNumber": "Renavam",
        "Color": "Cor",
        "FederalState": "Estado do veículo",
        "City": "Cidade",
        "ModelYear": "Ano do Modelo",
        "FabricationYear": "Ano de fabricação",
        "Fuel": "Combustível",
        "QRCode": "QRCode",
        "FipeCode": "Número da Fipe",
        "IdMapIcon": "Ícone do mapa",
        "IdMapIconColor": "Cor do Ícone",
        "ClientVehicleIntegrationCode": "Código de integração veículo do cliente",
        "Tracker1.TrackerIntegrationCode": "Código de integração Rastreador",
        "Tracker1.IdTracker": "ID do rastreador",
        "Tracker1.TrackerTemplateIntegrationCode": "Template do rastreador",
        "Tracker1.TrackerIMEI": "Imei do Rastreador",
        "Tracker1.Simcard1.ICCID": "ICCID do Chip",
        "Tracker1.Simcard1.APN": "APN do Chip",
        "Tracker1.Simcard1.CountryCode": "DDI",
        "Tracker1.Simcard1.AreaCode": "DDD",
        "Tracker1.Simcard1.PhoneNumber": "Número da linha",
    },
    "uo": {},
    "usuario": {},
}

LISTAGENS = {
    "clientes": dict(
        endpoint="/Administration/Client/List",
        payload=[{"PropertyName": "Active", "Condition": "Equal", "Value": True}],
        headers=["COD", "NOME", "DOC", "LOGIN"],
        campos=["ClientIntegrationCode", "TradingName", "DocumentNumber", "Login"],
    ),
    "veiculos": dict(
        endpoint="/Administration/Vehicle/List",
        payload=None,
        headers=["CLIENTE", "IDENTIFICAÇÃO", "ID DO RASTREADOR", "MODELO DO RASTREADOR", "TELEFONE DO CHIP"],
        campos=["ClientTradingName", "Identification", "IdTracker", "IdModelTracker", "PhoneNumber"],
    ),
    "rastreadores": dict(
        endpoint="/Administration/Tracker/List",
        payload=None,
        headers=["ID", "COD", "MODELO"],
        campos=["IdTracker", "TrackerIntegrationCode", "IdTrackerModel"],
    ),
    "pessoas": dict(
        endpoint="/Administration/Person/ListPerson",
        payload=None,
        headers=["CLIENTE", "LOGIN", "ATIVO"],
        campos=["ClientName", "Login", "IsActivatedClient"],
    ),
    "desatualizados": dict(
        endpoint="/Administration/Tracker/ListOutdatedTrackedUnits",
        payload=[{"PropertyName": "EventDate", "Condition": "Equal", "Value": None}],
        headers=["CLIENTE", "TRACKER", "UNIT"],
        campos=["ClientTradingName", "Tracker", "TrackedUnit"],
    ),
    "uos": dict(
        endpoint="/Administration/OrganizationalUnit/List",
        payload=None,
        headers=["NOME", "COD", "TIPO"],
        campos=["Name", "OrganizationalUnitIntegrationCode", "TypeOrganizationalUnit"],
    ),
}

# Armazenamento em memória dos Excel enviados (uso local, um usuário por vez)
UPLOADS = {}
UPLOADS_LOCK = threading.Lock()


def tratar_valor(caminho_completo, valor):
    """Trata o valor de acordo com o tipo esperado pelo campo (caminho completo, dot notation)."""
    v_str = str(valor).strip()
    if v_str.lower() == "nan" or v_str == "":
        return None
    if caminho_completo in INT_FIELDS:
        try:
            return int(float(v_str))
        except ValueError:
            return 0
    if caminho_completo in BOOL_FIELDS:
        return v_str.lower() in ["true", "1", "sim", "yes", "ativo"]
    return v_str


def set_nested_value(d, keys, value, caminho_completo):
    """Define valores em dicionários aninhados usando dot notation."""
    for key in keys[:-1]:
        d = d.setdefault(key, {})
    val_tratado = tratar_valor(caminho_completo, value)
    if val_tratado is not None:
        d[keys[-1]] = val_tratado


def montar_payload(tipo_config, mapping, row):
    payload = dict(tipo_config.get("defaults", {}))
    for param, col_excel in mapping.items():
        valor = row[col_excel]
        if "." in param:
            set_nested_value(payload, param.split("."), valor, param)
        else:
            val = tratar_valor(param, valor)
            if val is not None:
                payload[param] = val
    return payload


def requisicao_padrao(endpoint, payload, token):
    headers = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}
    r = requests.post(f"{BASE_URL}{endpoint}", json=payload if payload is not None else [], headers=headers, timeout=20)
    r.raise_for_status()
    return r.json()


@app.route("/")
def index():
    return render_template(
        "index.html",
        app_version=APP_VERSION,
        app_usuario=session.get("app_usuario_nome"),
        app_usuario_admin=session.get("app_usuario_admin", False),
    )


@app.route("/api/status")
def status():
    return jsonify(authenticated=bool(session.get("token")))


@app.route("/api/login", methods=["POST"])
def login():
    data = request.get_json(force=True) or {}
    usuario = str(data.get("login", "")).strip()
    senha = str(data.get("senha", ""))
    if not usuario or not senha:
        return jsonify(ok=False, error="Informe login e senha."), 400
    try:
        r = requests.post(f"{BASE_URL}/Login", data={"Username": usuario, "Password": senha}, timeout=10)
        r.raise_for_status()
        token = r.json().get("AccessToken")
        if not token:
            return jsonify(ok=False, error="Resposta sem token de acesso."), 400
        session["token"] = token
        return jsonify(ok=True)
    except requests.exceptions.RequestException as e:
        return jsonify(ok=False, error=f"Falha ao autenticar: {e}"), 400


@app.route("/api/logout", methods=["POST"])
def logout():
    session.clear()
    return jsonify(ok=True)


# --- LOGINS SALVOS (para trocar de cliente rapidamente) ---
# Guardados no Firestore (coleção "credenciais"), sobrevive a redeploys.
@app.route("/api/credenciais", methods=["GET"])
def listar_credenciais():
    docs = db.collection(CREDENCIAIS_COLLECTION).stream()
    lista = [dict(d.to_dict(), id=d.id) for d in docs]
    return jsonify(ok=True, credenciais=lista)


@app.route("/api/credenciais", methods=["POST"])
def criar_credencial():
    data = request.get_json(force=True) or {}
    nome = str(data.get("nome", "")).strip()
    login = str(data.get("login", "")).strip()
    senha = str(data.get("senha", ""))
    if not nome or not login or not senha:
        return jsonify(ok=False, error="Informe nome, login e senha."), 400
    doc_ref = db.collection(CREDENCIAIS_COLLECTION).document()
    dados = {"nome": nome, "login": login, "senha": senha}
    doc_ref.set(dados)
    return jsonify(ok=True, credencial=dict(dados, id=doc_ref.id))


@app.route("/api/credenciais/<cred_id>", methods=["PUT"])
def editar_credencial(cred_id):
    data = request.get_json(force=True) or {}
    nome = str(data.get("nome", "")).strip()
    login = str(data.get("login", "")).strip()
    senha = str(data.get("senha", ""))
    if not nome or not login or not senha:
        return jsonify(ok=False, error="Informe nome, login e senha."), 400
    doc_ref = db.collection(CREDENCIAIS_COLLECTION).document(cred_id)
    if not doc_ref.get().exists:
        return jsonify(ok=False, error="Login não encontrado."), 404
    dados = {"nome": nome, "login": login, "senha": senha}
    doc_ref.set(dados)
    return jsonify(ok=True, credencial=dict(dados, id=cred_id))


@app.route("/api/credenciais/<cred_id>", methods=["DELETE"])
def excluir_credencial(cred_id):
    doc_ref = db.collection(CREDENCIAIS_COLLECTION).document(cred_id)
    if not doc_ref.get().exists:
        return jsonify(ok=False, error="Login não encontrado."), 404
    doc_ref.delete()
    return jsonify(ok=True)


# --- COMANDOS DE RASTREADORES (Ferramentas > Comandos) ---
# Catálogo próprio, guardado no Firestore: uma coleção de modelos de rastreador
# e outra de comandos, cada comando amarrado a um modelo pelo campo modeloId.
COMANDO_MODELOS_COLLECTION = "comando_modelos"
COMANDO_ITENS_COLLECTION = "comando_itens"


@app.route("/api/comando-modelos", methods=["GET"])
def listar_comando_modelos():
    docs = db.collection(COMANDO_MODELOS_COLLECTION).stream()
    lista = [dict(d.to_dict(), id=d.id) for d in docs]
    return jsonify(ok=True, modelos=lista)


@app.route("/api/comando-modelos", methods=["POST"])
def criar_comando_modelo():
    data = request.get_json(force=True) or {}
    nome = str(data.get("nome", "")).strip()
    if not nome:
        return jsonify(ok=False, error="Informe o nome do modelo."), 400
    doc_ref = db.collection(COMANDO_MODELOS_COLLECTION).document()
    dados = {"nome": nome}
    doc_ref.set(dados)
    return jsonify(ok=True, modelo=dict(dados, id=doc_ref.id))


@app.route("/api/comando-modelos/<modelo_id>", methods=["PUT"])
def editar_comando_modelo(modelo_id):
    data = request.get_json(force=True) or {}
    nome = str(data.get("nome", "")).strip()
    if not nome:
        return jsonify(ok=False, error="Informe o nome do modelo."), 400
    doc_ref = db.collection(COMANDO_MODELOS_COLLECTION).document(modelo_id)
    if not doc_ref.get().exists:
        return jsonify(ok=False, error="Modelo não encontrado."), 404
    dados = {"nome": nome}
    doc_ref.set(dados)
    return jsonify(ok=True, modelo=dict(dados, id=modelo_id))


@app.route("/api/comando-modelos/<modelo_id>", methods=["DELETE"])
def excluir_comando_modelo(modelo_id):
    doc_ref = db.collection(COMANDO_MODELOS_COLLECTION).document(modelo_id)
    if not doc_ref.get().exists:
        return jsonify(ok=False, error="Modelo não encontrado."), 404
    # Apaga também os comandos cadastrados para esse modelo, pra não deixar órfãos.
    itens = db.collection(COMANDO_ITENS_COLLECTION).where(
        filter=FieldFilter("modeloId", "==", modelo_id)
    ).stream()
    for item in itens:
        item.reference.delete()
    doc_ref.delete()
    return jsonify(ok=True)


@app.route("/api/comando-itens", methods=["GET"])
def listar_comando_itens():
    docs = db.collection(COMANDO_ITENS_COLLECTION).stream()
    lista = [dict(d.to_dict(), id=d.id) for d in docs]
    return jsonify(ok=True, itens=lista)


@app.route("/api/comando-itens", methods=["POST"])
def criar_comando_item():
    data = request.get_json(force=True) or {}
    modelo_id = str(data.get("modeloId", "")).strip()
    nome = str(data.get("nome", "")).strip()
    comando = str(data.get("comando", "")).strip()
    if not modelo_id or not nome or not comando:
        return jsonify(ok=False, error="Informe modelo, nome e comando."), 400
    doc_ref = db.collection(COMANDO_ITENS_COLLECTION).document()
    dados = {"modeloId": modelo_id, "nome": nome, "comando": comando}
    doc_ref.set(dados)
    return jsonify(ok=True, item=dict(dados, id=doc_ref.id))


@app.route("/api/comando-itens/<item_id>", methods=["PUT"])
def editar_comando_item(item_id):
    data = request.get_json(force=True) or {}
    modelo_id = str(data.get("modeloId", "")).strip()
    nome = str(data.get("nome", "")).strip()
    comando = str(data.get("comando", "")).strip()
    if not modelo_id or not nome or not comando:
        return jsonify(ok=False, error="Informe modelo, nome e comando."), 400
    doc_ref = db.collection(COMANDO_ITENS_COLLECTION).document(item_id)
    if not doc_ref.get().exists:
        return jsonify(ok=False, error="Comando não encontrado."), 404
    dados = {"modeloId": modelo_id, "nome": nome, "comando": comando}
    doc_ref.set(dados)
    return jsonify(ok=True, item=dict(dados, id=item_id))


@app.route("/api/comando-itens/<item_id>", methods=["DELETE"])
def excluir_comando_item(item_id):
    doc_ref = db.collection(COMANDO_ITENS_COLLECTION).document(item_id)
    if not doc_ref.get().exists:
        return jsonify(ok=False, error="Comando não encontrado."), 404
    doc_ref.delete()
    return jsonify(ok=True)


# --- CLIENTES EM IMPLANTAÇÃO (menu "Implantação") ---
# Cadastro simples e independente do resto do app: acompanha clientes que estão
# entrando na base (data de entrada, objetivo, valor do contrato, CSM responsável).
# Cada cliente tem uma subcoleção "eventos" (a linha do tempo por setor); a
# "última ação" exibida na listagem é sempre calculada a partir dela, nunca
# digitada — é só o acontecimento mais recente entre os 4 setores.
IMPLANTACAO_CLIENTES_COLLECTION = "implantacao_clientes"
IMPLANTACAO_SETORES = ["Implantação", "Migração", "Suporte", "Comercial"]


def _dados_implantacao_cliente(data):
    cliente = str(data.get("cliente", "")).strip()
    if not cliente:
        return None
    try:
        valor_contrato = float(data.get("valor_contrato") or 0)
    except (TypeError, ValueError):
        valor_contrato = 0
    return {
        "cliente": cliente,
        "data_entrada": str(data.get("data_entrada", "")).strip(),
        "objetivo": str(data.get("objetivo", "")).strip(),
        "valor_contrato": valor_contrato,
        "csm": str(data.get("csm", "")).strip(),
    }


def _dados_implantacao_evento(data):
    setor = str(data.get("setor", "")).strip()
    titulo = str(data.get("titulo", "")).strip()
    if not titulo or setor not in IMPLANTACAO_SETORES:
        return None
    return {
        "setor": setor,
        "titulo": titulo,
        "descricao": str(data.get("descricao", "")).strip(),
        "data": str(data.get("data", "")).strip(),
        "responsavel": str(data.get("responsavel", "")).strip(),
    }


@app.route("/api/implantacao/clientes", methods=["GET"])
def listar_implantacao_clientes():
    docs = list(db.collection(IMPLANTACAO_CLIENTES_COLLECTION).stream())
    lista = []
    for d in docs:
        cliente = dict(d.to_dict(), id=d.id)
        ultimo = list(
            d.reference.collection("eventos")
            .order_by("data", direction=firestore.Query.DESCENDING)
            .limit(1)
            .stream()
        )
        if ultimo:
            ev = ultimo[0].to_dict()
            cliente["ultima_acao"] = f"[{ev.get('setor')}] {ev.get('titulo')}"
            cliente["ultima_acao_data"] = ev.get("data") or ""
        else:
            cliente["ultima_acao"] = ""
            cliente["ultima_acao_data"] = ""
        lista.append(cliente)

    # Mais recente primeiro; "" (sem eventos ainda) sempre por último.
    lista.sort(key=lambda c: c["ultima_acao_data"], reverse=True)
    return jsonify(ok=True, clientes=lista)


@app.route("/api/implantacao/clientes", methods=["POST"])
def criar_implantacao_cliente():
    data = request.get_json(force=True) or {}
    dados = _dados_implantacao_cliente(data)
    if dados is None:
        return jsonify(ok=False, error="Informe o nome do cliente."), 400
    doc_ref = db.collection(IMPLANTACAO_CLIENTES_COLLECTION).document()
    doc_ref.set(dados)
    return jsonify(ok=True, cliente=dict(dados, id=doc_ref.id))


@app.route("/api/implantacao/clientes/<cliente_id>", methods=["PUT"])
def editar_implantacao_cliente(cliente_id):
    data = request.get_json(force=True) or {}
    dados = _dados_implantacao_cliente(data)
    if dados is None:
        return jsonify(ok=False, error="Informe o nome do cliente."), 400
    doc_ref = db.collection(IMPLANTACAO_CLIENTES_COLLECTION).document(cliente_id)
    if not doc_ref.get().exists:
        return jsonify(ok=False, error="Cliente não encontrado."), 404
    doc_ref.set(dados)
    return jsonify(ok=True, cliente=dict(dados, id=cliente_id))


@app.route("/api/implantacao/clientes/<cliente_id>", methods=["DELETE"])
def excluir_implantacao_cliente(cliente_id):
    doc_ref = db.collection(IMPLANTACAO_CLIENTES_COLLECTION).document(cliente_id)
    if not doc_ref.get().exists:
        return jsonify(ok=False, error="Cliente não encontrado."), 404
    for ev in doc_ref.collection("eventos").stream():
        ev.reference.delete()
    doc_ref.delete()
    return jsonify(ok=True)


@app.route("/api/implantacao/clientes/<cliente_id>/eventos", methods=["GET"])
def listar_implantacao_eventos(cliente_id):
    doc_ref = db.collection(IMPLANTACAO_CLIENTES_COLLECTION).document(cliente_id)
    if not doc_ref.get().exists:
        return jsonify(ok=False, error="Cliente não encontrado."), 404
    docs = doc_ref.collection("eventos").stream()
    lista = [dict(d.to_dict(), id=d.id) for d in docs]
    return jsonify(ok=True, eventos=lista)


@app.route("/api/implantacao/clientes/<cliente_id>/eventos", methods=["POST"])
def criar_implantacao_evento(cliente_id):
    doc_ref = db.collection(IMPLANTACAO_CLIENTES_COLLECTION).document(cliente_id)
    if not doc_ref.get().exists:
        return jsonify(ok=False, error="Cliente não encontrado."), 404
    data = request.get_json(force=True) or {}
    dados = _dados_implantacao_evento(data)
    if dados is None:
        return jsonify(ok=False, error="Informe o setor e o título."), 400
    evento_ref = doc_ref.collection("eventos").document()
    evento_ref.set(dados)
    return jsonify(ok=True, evento=dict(dados, id=evento_ref.id))


@app.route("/api/implantacao/clientes/<cliente_id>/eventos/<evento_id>", methods=["PUT"])
def editar_implantacao_evento(cliente_id, evento_id):
    data = request.get_json(force=True) or {}
    dados = _dados_implantacao_evento(data)
    if dados is None:
        return jsonify(ok=False, error="Informe o setor e o título."), 400
    evento_ref = db.collection(IMPLANTACAO_CLIENTES_COLLECTION).document(cliente_id).collection("eventos").document(evento_id)
    if not evento_ref.get().exists:
        return jsonify(ok=False, error="Acontecimento não encontrado."), 404
    evento_ref.set(dados)
    return jsonify(ok=True, evento=dict(dados, id=evento_id))


@app.route("/api/implantacao/clientes/<cliente_id>/eventos/<evento_id>", methods=["DELETE"])
def excluir_implantacao_evento(cliente_id, evento_id):
    evento_ref = db.collection(IMPLANTACAO_CLIENTES_COLLECTION).document(cliente_id).collection("eventos").document(evento_id)
    if not evento_ref.get().exists:
        return jsonify(ok=False, error="Acontecimento não encontrado."), 404
    evento_ref.delete()
    return jsonify(ok=True)


# --- CLIENTES EM MIGRAÇÃO ---
# Coleção independente: uma linha nasce quando alguém importa veículos com a
# opção "Criar planilha" marcada (não mais atrelada aos logins salvos).
MIGRACAO_COLLECTION = "migracao_clientes"
CAMPOS_MIGRACAO_PADRAO = {
    "nome": "",
    "cs": "",
    "plataforma_origem": "",
    "qtd_clientes": 0,
    "qtd_placas": 0,
    "percentual_migracao": 0,
}


def _para_int(valor):
    try:
        return int(float(valor))
    except (TypeError, ValueError):
        return 0


def _para_float(valor):
    try:
        return float(valor)
    except (TypeError, ValueError):
        return 0.0


def obter_ou_criar_cliente_migracao(nome):
    """Retorna o id da linha do cliente (por nome), criando-a se ainda não existir."""
    query = db.collection(MIGRACAO_COLLECTION).where(filter=FieldFilter("nome", "==", nome)).limit(1).stream()
    existente = next(query, None)
    if existente:
        return existente.id
    dados = dict(CAMPOS_MIGRACAO_PADRAO)
    dados["nome"] = nome
    doc_ref = db.collection(MIGRACAO_COLLECTION).document()
    doc_ref.set(dados)
    return doc_ref.id


def _veiculo_doc_id(cliente, veiculo):
    """ID determinístico por (cliente, veículo) — reimportar o mesmo veículo atualiza a
    linha existente em vez de duplicar."""
    chave = f"{cliente}||{veiculo}".encode("utf-8")
    return hashlib.sha1(chave).hexdigest()


STATUS_VEICULO_VALIDOS = ["Aguardando", "Enviado", "Migrado", "Enviar"]
STATUS_VEICULO_PADRAO = "Aguardando"


def salvar_veiculos_migracao(cliente_migracao_id, veiculos):
    """Upsert por (cliente, veículo). Usa merge para não apagar 'comando'/'status' já definidos.
    Veículos novos entram com status "Aguardando"; veículos reimportados mantêm o status atual."""
    subcolecao = db.collection(MIGRACAO_COLLECTION).document(cliente_migracao_id).collection("veiculos")
    for v in veiculos:
        doc_id = _veiculo_doc_id(v["cliente"], v["veiculo"])
        doc_ref = subcolecao.document(doc_id)
        dados = dict(v)
        if not doc_ref.get().exists:
            dados["status"] = STATUS_VEICULO_PADRAO
        doc_ref.set(dados, merge=True)


def recalcular_contagens_migracao(cliente_migracao_id):
    """Recalcula qtd_clientes/qtd_placas a partir do total acumulado de veículos salvos."""
    subcolecao = db.collection(MIGRACAO_COLLECTION).document(cliente_migracao_id).collection("veiculos")
    docs = [d.to_dict() for d in subcolecao.stream()]
    clientes = {d.get("cliente") for d in docs if d.get("cliente")}
    placas = {d.get("veiculo") for d in docs if d.get("veiculo")}
    qtd_clientes, qtd_placas = len(clientes), len(placas)
    db.collection(MIGRACAO_COLLECTION).document(cliente_migracao_id).update({
        "qtd_clientes": qtd_clientes, "qtd_placas": qtd_placas,
    })
    return qtd_clientes, qtd_placas


@app.route("/api/migracao/clientes", methods=["GET"])
def listar_clientes_migracao():
    docs = db.collection(MIGRACAO_COLLECTION).stream()
    lista = [dict(d.to_dict(), id=d.id) for d in docs]
    lista.sort(key=lambda c: c["nome"].lower())
    return jsonify(ok=True, clientes=lista)


@app.route("/api/migracao/clientes", methods=["POST"])
def criar_cliente_migracao():
    data = request.get_json(force=True) or {}
    nome = str(data.get("nome", "")).strip()
    if not nome:
        return jsonify(ok=False, error="Informe o nome do cliente."), 400
    cliente_id = obter_ou_criar_cliente_migracao(nome)
    doc = db.collection(MIGRACAO_COLLECTION).document(cliente_id).get()
    return jsonify(ok=True, cliente=dict(doc.to_dict(), id=cliente_id))


@app.route("/api/migracao/clientes/<cliente_id>", methods=["DELETE"])
def excluir_cliente_migracao(cliente_id):
    doc_ref = db.collection(MIGRACAO_COLLECTION).document(cliente_id)
    if not doc_ref.get().exists:
        return jsonify(ok=False, error="Cliente não encontrado."), 404
    for v in doc_ref.collection("veiculos").stream():
        v.reference.delete()
    doc_ref.delete()
    return jsonify(ok=True)


@app.route("/api/migracao/clientes/<cliente_id>", methods=["PUT"])
def atualizar_cliente_migracao(cliente_id):
    doc_ref = db.collection(MIGRACAO_COLLECTION).document(cliente_id)
    atual = doc_ref.get()
    if not atual.exists:
        return jsonify(ok=False, error="Cliente não encontrado."), 404
    data = request.get_json(force=True) or {}
    dados = {
        "cs": str(data.get("cs", "")).strip(),
        "plataforma_origem": str(data.get("plataforma_origem", "")).strip(),
        "qtd_clientes": _para_int(data.get("qtd_clientes")),
        "qtd_placas": _para_int(data.get("qtd_placas")),
        "percentual_migracao": _para_float(data.get("percentual_migracao")),
    }
    doc_ref.update(dados)
    return jsonify(ok=True, dados=dict(atual.to_dict(), **dados))


@app.route("/api/migracao/clientes/<cliente_id>/veiculos", methods=["GET"])
def listar_veiculos_migracao(cliente_id):
    if not db.collection(MIGRACAO_COLLECTION).document(cliente_id).get().exists:
        return jsonify(ok=False, error="Cliente não encontrado."), 404
    docs = db.collection(MIGRACAO_COLLECTION).document(cliente_id).collection("veiculos").stream()
    lista = []
    for d in docs:
        dados = d.to_dict()
        dados.setdefault("comando", "")
        dados.setdefault("status", STATUS_VEICULO_PADRAO)
        lista.append(dict(dados, id=d.id))
    lista.sort(key=lambda v: (v.get("cliente") or "", v.get("veiculo") or ""))
    return jsonify(ok=True, veiculos=lista)


@app.route("/api/migracao/clientes/<cliente_id>/veiculos/<veiculo_id>", methods=["PUT"])
def atualizar_veiculo_migracao(cliente_id, veiculo_id):
    ref = db.collection(MIGRACAO_COLLECTION).document(cliente_id).collection("veiculos").document(veiculo_id)
    if not ref.get().exists:
        return jsonify(ok=False, error="Veículo não encontrado."), 404
    data = request.get_json(force=True) or {}
    atualizacoes = {}
    if "comando" in data:
        atualizacoes["comando"] = str(data.get("comando", ""))
    if "status" in data:
        status = str(data.get("status", "")).strip()
        if status not in STATUS_VEICULO_VALIDOS:
            return jsonify(ok=False, error="Status inválido."), 400
        atualizacoes["status"] = status
    if not atualizacoes:
        return jsonify(ok=False, error="Nada para atualizar."), 400
    ref.update(atualizacoes)
    return jsonify(ok=True, **atualizacoes)


CAMPOS_VEICULO_EDITAVEIS = ["cliente", "veiculo", "equipamento", "id_equipamento", "numero_linha", "comando"]


@app.route("/api/migracao/clientes/<cliente_id>/veiculos/salvar-lote", methods=["POST"])
def salvar_lote_veiculos_migracao(cliente_id):
    if not db.collection(MIGRACAO_COLLECTION).document(cliente_id).get().exists:
        return jsonify(ok=False, error="Cliente não encontrado."), 404
    body = request.get_json(force=True) or {}
    itens = body.get("veiculos") or []
    subcolecao = db.collection(MIGRACAO_COLLECTION).document(cliente_id).collection("veiculos")

    for item in itens:
        cliente = str(item.get("cliente", "")).strip()
        veiculo = str(item.get("veiculo", "")).strip()
        if not cliente or not veiculo:
            continue  # linha em branco (não preenchida) — ignora

        novo_id = _veiculo_doc_id(cliente, veiculo)
        antigo_id = item.get("id")
        dados = {campo: str(item.get(campo, "")) for campo in CAMPOS_VEICULO_EDITAVEIS}

        novo_ref = subcolecao.document(novo_id)
        if antigo_id and antigo_id != novo_id:
            # Cliente/Veículo mudaram: era um registro existente (id real) —
            # migra o status pro novo doc e remove o antigo pra não duplicar.
            antigo_ref = subcolecao.document(antigo_id)
            antigo_doc = antigo_ref.get()
            if antigo_doc.exists:
                status_existente = antigo_doc.to_dict().get("status")
                if status_existente:
                    dados["status"] = status_existente
                antigo_ref.delete()
        if "status" not in dados and not novo_ref.get().exists:
            dados["status"] = STATUS_VEICULO_PADRAO

        novo_ref.set(dados, merge=True)

    qtd_clientes, qtd_placas = recalcular_contagens_migracao(cliente_id)
    return jsonify(ok=True, qtd_clientes=qtd_clientes, qtd_placas=qtd_placas)


@app.route("/api/dashboard")
def dashboard_indicadores():
    total_clientes = sum(1 for _ in db.collection(MIGRACAO_COLLECTION).stream())

    por_status = {s: 0 for s in STATUS_VEICULO_VALIDOS}
    total_veiculos = 0
    for doc in db.collection_group("veiculos").stream():
        total_veiculos += 1
        status = doc.to_dict().get("status") or STATUS_VEICULO_PADRAO
        por_status[status] = por_status.get(status, 0) + 1

    return jsonify(
        ok=True,
        total_clientes=total_clientes,
        total_veiculos=total_veiculos,
        por_status=por_status,
    )


@app.route("/api/list/<tipo>")
def listar(tipo):
    token = session.get("token")
    if not token:
        return jsonify(ok=False, error="Autentique-se primeiro."), 401
    config = LISTAGENS.get(tipo)
    if not config:
        return jsonify(ok=False, error="Listagem desconhecida."), 404
    try:
        data = requisicao_padrao(config["endpoint"], config["payload"], token)
    except requests.exceptions.RequestException as e:
        return jsonify(ok=False, error=f"Falha na consulta: {e}"), 400
    rows = [[item.get(c) for c in config["campos"]] for item in (data or [])]
    return jsonify(ok=True, headers=config["headers"], rows=rows)


# Exportação genérica: recebe o headers/rows já carregado em tela (qualquer uma
# das Consultas/Listagens) e devolve como planilha, sem precisar consultar a SSX de novo.
@app.route("/api/exportar-excel", methods=["POST"])
def exportar_excel():
    data = request.get_json(force=True) or {}
    headers = data.get("headers") or []
    rows = data.get("rows") or []
    nome = "".join(c for c in str(data.get("nome") or "") if c.isalnum() or c in "_-") or "exportacao"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(headers)
    for row in rows:
        ws.append(row)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return Response(
        buffer.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{nome}.xlsx"'},
    )


@app.route("/api/import/params/<tipo>")
def import_params(tipo):
    config = PARAMS.get(tipo)
    if not config:
        return jsonify(ok=False, error="Tipo desconhecido."), 404
    labels = LABELS.get(tipo, {})
    campos = [{"nome": c, "rotulo": labels.get(c, c)} for c in config["campos"]]
    return jsonify(ok=True, titulo=config["titulo"], campos=campos)


@app.route("/api/import/upload", methods=["POST"])
def import_upload():
    if "arquivo" not in request.files:
        return jsonify(ok=False, error="Nenhum arquivo enviado."), 400
    arquivo = request.files["arquivo"]
    try:
        df = pd.read_excel(arquivo, dtype=str)
    except Exception as e:
        return jsonify(ok=False, error=f"Falha ao ler Excel: {e}"), 400
    file_id = uuid.uuid4().hex
    with UPLOADS_LOCK:
        UPLOADS[file_id] = df
    return jsonify(ok=True, file_id=file_id, colunas=list(df.columns), total_linhas=len(df))


# Importação roda em thread separada, pelo mesmo motivo do envio de comando em
# massa: uma importação grande facilmente passa dos ~30s que o Render/gunicorn
# tolera num único request parado, cortando a importação no meio.
IMPORT_JOBS = {}
IMPORT_JOBS_LOCK = threading.Lock()


def _extrair_aninhado(payload, *chaves):
    atual = payload
    for chave in chaves:
        if not isinstance(atual, dict):
            return None
        atual = atual.get(chave)
    return atual


def _executar_import(job_id, tipo_config, mapping, df, endpoint, token, criar_planilha, nome_cliente_planilha):
    sucessos = erros = 0
    veiculos_para_planilha = []
    headers = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}

    for pos, (index, row) in enumerate(df.iterrows()):
        payload = montar_payload(tipo_config, mapping, row)
        try:
            resp = requests.post(f"{BASE_URL}{endpoint}", json=payload, headers=headers, timeout=20)
            if resp.status_code in (200, 201):
                sucessos += 1
                if criar_planilha:
                    cliente = payload.get("ClientIntegrationCode")
                    veiculo = payload.get("Identification")
                    if cliente and veiculo:
                        ddi = _extrair_aninhado(payload, "Tracker1", "Simcard1", "CountryCode") or ""
                        ddd = _extrair_aninhado(payload, "Tracker1", "Simcard1", "AreaCode") or ""
                        numero = _extrair_aninhado(payload, "Tracker1", "Simcard1", "PhoneNumber") or ""
                        equipamento = _extrair_aninhado(payload, "Tracker1", "TrackerTemplateIntegrationCode") or ""
                        id_equipamento = _extrair_aninhado(payload, "Tracker1", "IdTracker") or ""
                        veiculos_para_planilha.append({
                            "cliente": str(cliente),
                            "veiculo": str(veiculo),
                            "equipamento": str(equipamento),
                            "id_equipamento": str(id_equipamento),
                            "numero_linha": f"{ddi}{ddd}{numero}",
                        })
            else:
                erros += 1
                with IMPORT_JOBS_LOCK:
                    IMPORT_JOBS[job_id]["logs"].append(f"Erro linha {pos + 1}: {resp.text}")
        except requests.exceptions.RequestException as e:
            erros += 1
            with IMPORT_JOBS_LOCK:
                IMPORT_JOBS[job_id]["logs"].append(f"Erro linha {pos + 1}: {e}")

        with IMPORT_JOBS_LOCK:
            job = IMPORT_JOBS[job_id]
            job["atual"] = pos + 1
            job["sucessos"] = sucessos
            job["erros"] = erros

    resultado_planilha = None
    if criar_planilha:
        cliente_migracao_id = obter_ou_criar_cliente_migracao(nome_cliente_planilha)
        if veiculos_para_planilha:
            salvar_veiculos_migracao(cliente_migracao_id, veiculos_para_planilha)
        qtd_clientes, qtd_placas = recalcular_contagens_migracao(cliente_migracao_id)
        resultado_planilha = {"nome": nome_cliente_planilha, "qtd_clientes": qtd_clientes, "qtd_placas": qtd_placas}

    with IMPORT_JOBS_LOCK:
        IMPORT_JOBS[job_id]["status"] = "concluido"
        IMPORT_JOBS[job_id]["planilha"] = resultado_planilha


@app.route("/api/import/run", methods=["POST"])
def import_run():
    token = session.get("token")
    if not token:
        return jsonify(ok=False, error="Autentique-se primeiro."), 401

    body = request.get_json(force=True) or {}
    tipo = body.get("tipo")
    file_id = body.get("file_id")
    mapping = body.get("mapping") or {}
    criar_planilha = bool(body.get("criar_planilha")) and tipo == "veiculo"
    nome_cliente_planilha = str(body.get("nome_cliente_planilha", "")).strip()

    tipo_config = PARAMS.get(tipo)
    if not tipo_config:
        return jsonify(ok=False, error="Tipo desconhecido."), 404
    if not mapping:
        return jsonify(ok=False, error="Mapeie ao menos uma coluna."), 400
    if criar_planilha and not nome_cliente_planilha:
        return jsonify(ok=False, error="Informe o nome do cliente para criar a planilha em Clientes em migração."), 400

    with UPLOADS_LOCK:
        df = UPLOADS.pop(file_id, None)
    if df is None:
        return jsonify(ok=False, error="Arquivo não encontrado. Envie novamente."), 400

    endpoint = tipo_config["endpoint"]
    total = len(df)

    job_id = uuid.uuid4().hex
    with IMPORT_JOBS_LOCK:
        IMPORT_JOBS[job_id] = {
            "status": "rodando", "atual": 0, "total": total,
            "sucessos": 0, "erros": 0, "logs": [], "planilha": None,
        }

    threading.Thread(
        target=_executar_import,
        args=(job_id, tipo_config, mapping, df, endpoint, token, criar_planilha, nome_cliente_planilha),
        daemon=True,
    ).start()

    return jsonify(ok=True, job_id=job_id, total=total)


@app.route("/api/import/run/status/<job_id>")
def import_run_status(job_id):
    with IMPORT_JOBS_LOCK:
        job = IMPORT_JOBS.get(job_id)
        if not job:
            return jsonify(ok=False, error="Job não encontrado."), 404
        resultado = dict(job)
    return jsonify(ok=True, **resultado)


# --- ENVIO DE COMANDO (SMS Market) ---
# Portado do "Configurador de rastreadores V5.0" (tkinter). A lógica de geração
# de comando por modelo/comando e o fluxo de envio (único, livre, em massa)
# foram mantidos fiéis ao programa original, inclusive peculiaridades dele
# (ex.: alguns modelos retornam um texto diferente do exibido na tela original;
# aqui exibimos sempre o texto que é realmente enviado).
SMS_BASE_URL = "https://api.smsmarket.com.br/webservice-rest"

UPLOADS_MASSA = {}
UPLOADS_MASSA_LOCK = threading.Lock()


def gerar_comando_sms(modelo, comando, id_, apn, loginapn, porta, operadora):
    """Retorna o texto do comando a ser enviado, ou None se a combinação não é suportada."""

    if modelo == "E3/E3+":
        if comando == "REG000000#":
            return "REG000000#"
        if comando == "SMS1":
            return "SMS1"
        if comando == "IP/Porta1":
            return f"IP1#200.152.62.20#{porta}#"
        if comando == "IP/Porta2":
            return f"IP2#200.152.62.20#{porta}#"
        if comando == "SMS0":
            return "SMS0"

    if modelo == "F1/M1":
        if comando == "IP/Porta":
            return f"SERVER,0,200.152.62.20,{porta},0#"
        if comando == "APN":
            return f"APN,{apn},{loginapn},{loginapn}#"
        if comando == "Reset":
            return "#reiniciar,888888#"

    if modelo == "ITR-120/155":
        if comando == "IP/Porta":
            return f"SERVER,0,200.152.62.20,{porta},0#"
        if comando == "APN":
            return f"APN,{apn},{loginapn},{loginapn}#"
        if comando == "Reset":
            return "RESET#"

    if modelo == "JC181":
        if comando == "COREKITSW,0":
            return "COREKITSW,0"
        if comando == "APN":
            return f"APN,{apn},{loginapn},{loginapn}"
        if comando == "URLTYPE,2":
            return "URLTYPE,2"
        if comando == "SERVER":
            return "SERVER,0,200.152.62.22,21122"

    if modelo == "JC450":
        if comando == "URLTYPE,2":
            return "URLTYPE,2"
        if comando == "APN":
            return f"APN,jimi,{apn},,,,,,{loginapn},,{loginapn},,,,,IP,IP,"
        if comando == "SERVER":
            return "SERVER,jimi.systemsatx.com.br,21122,NA,NA,NA,NA"
        if comando == "LOCATEREP":
            return "LOCATEREP,60"
        if comando == "SHUTDOWNTIME":
            return "SHUTDOWNTIME,120"
        if comando == "WAKEMODE":
            return "WAKEMODE,103"

    if modelo in ("VL01/02/03", "LV12", "N4", "J16"):
        if comando == "IP/Porta":
            return f"SERVER,0,200.152.62.20,{porta},0#"
        if comando == "APN":
            return f"APN,{apn},{loginapn},{loginapn}#"
        if comando == "Reset":
            return "RESET#"

    if modelo == "NT20":
        if comando == "IP/Porta":
            return f"SERVER,8520,200.152.62.20,{porta},0#"
        if comando == "APN":
            return f"APN,{apn},{loginapn},{loginapn}#"
        if comando == "Reset":
            return "RESET#"

    if modelo in ("ST40XX", "ST80XX"):
        if comando == "IP/Porta":
            return f"PRG;{id_};10;05#200.152.62.20;06#{porta};08#200.152.62.20;09#{porta}"
        if comando == "APN":
            return f"PRG;{apn};10;00#01;01#{apn};02#{loginapn};03#{loginapn}"
        if comando == "Rede zip":
            return f"PRG;{apn};10;00#01;01#{apn};02#{loginapn};03#{loginapn}"
        if comando == "IG Física":
            return f"PRG;{id_};17;00#01"
        if comando == "Reset":
            return f"CMD;{id_};03;03"

    if modelo == "ST3XX":
        if comando == "IP/Porta":
            if operadora == "Outras":
                return f"ST300NTW;{id_};02;1;{apn};{loginapn};{loginapn};200.152.62.20;{porta};200.152.62.20;{porta};;"
            return f"ST300NTW;{id_};02;0;{apn};{loginapn};{loginapn};200.152.62.20;{porta};200.152.62.20;{porta};;"
        if comando == "Rede zip":
            return f"ST300NTW;{id_};02;1;{apn};{loginapn};{loginapn};200.152.62.20;{porta};200.152.62.20;{porta};;"
        if comando == "Reset":
            return f"ST300CMD;{id_};02;Reboot"

    if modelo == "TK311":
        if comando == "IP/Porta":
            return f"adminip123456 200.152.62.20 {porta}"
        if comando == "Reset":
            return "reset123456"

    if modelo == "JC400AD":
        if comando == "COREKITSW":
            return "COREKITSW,0"
        if comando == "APN":
            return f"APN,SYSTEMSAT,{apn},,,,,,{loginapn},,{loginapn},,,,,IPv4,IPv4,,"
        if comando == "SERVER":
            return "SERVER#1#jimi.systemsatx.com.br#21100"
        if comando == "RSERVICE":
            return "RSERVICE,jimi.systemsatx.com.br:1936/live"
        if comando == "UPLOAD":
            return "UPLOAD,http://jimi.systemsatx.com.br:23010/upload"
        if comando == "FILELIST":
            return "FILELIST,https://jimiapi.systemsatx.com.br/fileList"
        if comando == "Reset":
            return "Reboot"

    if modelo in ("GTK LW", "TR05"):
        if comando == "IP/Porta":
            return f"SERVER,8888,200.152.62.20,{porta}#"
        if comando == "APN":
            return f"APN,{apn},{loginapn},{loginapn}#"
        if comando == "Reset":
            return "RESET#"

    return None


def _consultar_saldo_sms(usuario, senha):
    r = requests.get(f"{SMS_BASE_URL}/balance", params={"user": usuario, "password": senha}, timeout=15)
    r.raise_for_status()
    return r.json().get("balance_2")


def _enviar_sms(numero, conteudo, campaign_id, auth=None):
    # auth explícito é usado pela thread de envio em massa, que roda fora do
    # contexto de requisição e por isso não tem acesso à `session`.
    auth = auth or session.get("sms_auth")
    if not auth:
        return None, "Autentique-se na SMS Market primeiro."
    headers = {"Authorization": f"Basic {auth}"}
    payload = {"number": numero, "content": conteudo, "type": "0", "campaign_id": campaign_id}
    try:
        r = requests.post(f"{SMS_BASE_URL}/send-single.php", data=payload, headers=headers, timeout=20)
        data = r.json()
        return data.get("responseDescription"), None
    except requests.exceptions.RequestException as e:
        return None, str(e)
    except ValueError:
        return None, "Resposta inválida da SMS Market."


@app.route("/api/comando/autenticar", methods=["POST"])
def comando_autenticar():
    data = request.get_json(force=True) or {}
    usuario = str(data.get("usuario", "")).strip()
    senha = str(data.get("senha", ""))
    if not usuario or not senha:
        return jsonify(ok=False, error="Informe usuário e senha."), 400
    try:
        saldo = _consultar_saldo_sms(usuario, senha)
    except requests.exceptions.RequestException as e:
        return jsonify(ok=False, error=f"Falha ao consultar saldo: {e}"), 400
    if saldo is None or saldo == "None":
        return jsonify(ok=False, error="Usuário ou senha inválido."), 400
    session["sms_usuario"] = usuario
    session["sms_senha"] = senha
    session["sms_auth"] = base64.b64encode(f"{usuario}:{senha}".encode()).decode()
    return jsonify(ok=True, saldo=saldo)


@app.route("/api/comando/saldo")
def comando_saldo():
    usuario, senha = session.get("sms_usuario"), session.get("sms_senha")
    if not usuario or not senha:
        return jsonify(ok=False, error="Autentique-se na SMS Market primeiro."), 401
    try:
        saldo = _consultar_saldo_sms(usuario, senha)
    except requests.exceptions.RequestException as e:
        return jsonify(ok=False, error=f"Falha ao consultar saldo: {e}"), 400
    return jsonify(ok=True, saldo=saldo)


@app.route("/api/comando/gerar", methods=["POST"])
def comando_gerar():
    data = request.get_json(force=True) or {}
    texto = gerar_comando_sms(
        modelo=data.get("modelo", ""),
        comando=data.get("comando", ""),
        id_=data.get("id", ""),
        apn=data.get("apn", ""),
        loginapn=data.get("loginapn", ""),
        porta=data.get("porta", ""),
        operadora=data.get("operadora", ""),
    )
    if texto is None:
        return jsonify(ok=False, error="Comando não implementado para este modelo."), 400
    return jsonify(ok=True, texto=texto)


@app.route("/api/comando/enviar", methods=["POST"])
def comando_enviar():
    if not session.get("sms_auth"):
        return jsonify(ok=False, error="Autentique-se na SMS Market primeiro."), 401
    data = request.get_json(force=True) or {}
    numero = str(data.get("numero", "")).strip()
    conteudo = str(data.get("conteudo", ""))
    campaign_id = str(data.get("campaign_id", "Envio de comando"))
    if not numero or not conteudo:
        return jsonify(ok=False, error="Informe número e conteúdo."), 400

    resposta, erro = _enviar_sms(numero, conteudo, campaign_id)
    if erro:
        return jsonify(ok=False, error=erro), 400

    saldo = None
    usuario, senha = session.get("sms_usuario"), session.get("sms_senha")
    if usuario and senha:
        try:
            saldo = _consultar_saldo_sms(usuario, senha)
        except requests.exceptions.RequestException:
            pass
    return jsonify(ok=True, resposta=resposta, saldo=saldo)


@app.route("/api/comando/upload-massa", methods=["POST"])
def comando_upload_massa():
    if "arquivo" not in request.files:
        return jsonify(ok=False, error="Nenhum arquivo enviado."), 400
    conteudo = request.files["arquivo"].read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(conteudo))
        # max_row inclui a linha 1 (cabeçalho), que também é enviada como SMS
        # (comportamento original preservado) — refletimos o total real de envios.
        total_linhas = max(wb.active.max_row, 0)
    except Exception as e:
        return jsonify(ok=False, error=f"Falha ao ler Excel: {e}"), 400
    file_id = uuid.uuid4().hex
    with UPLOADS_MASSA_LOCK:
        UPLOADS_MASSA[file_id] = conteudo
    return jsonify(ok=True, file_id=file_id, total_linhas=total_linhas)


# Envio em massa roda em thread separada (não presa a um único request HTTP),
# porque o Render/gunicorn mata requisições longas paradas (timeout padrão de
# 30s) — com muitas linhas e intervalo entre SMS's isso facilmente estoura.
# A tela consulta o andamento via polling em /status/<job_id>.
COMANDO_JOBS = {}
COMANDO_JOBS_LOCK = threading.Lock()


def _executar_envio_massa(job_id, coluna1, coluna2, coluna3, intervalo, auth, usuario, senha):
    sucessos = erros = 0
    linha = 1
    total = max(len(coluna1), 1)
    valor_coluna1 = coluna1[1] if len(coluna1) > 1 else "None"

    while valor_coluna1 != "None":
        valor_coluna1 = coluna1[linha - 1] if linha <= len(coluna1) else None
        valor_coluna2 = coluna2[linha - 1] if linha <= len(coluna2) else None
        valor_coluna3 = coluna3[linha - 1] if linha <= len(coluna3) else None

        if valor_coluna1 is None and valor_coluna2 is None:
            break

        resposta, erro = _enviar_sms(valor_coluna2, valor_coluna1, valor_coluna3, auth=auth)
        with COMANDO_JOBS_LOCK:
            job = COMANDO_JOBS[job_id]
            if erro:
                erros += 1
                job["logs"].append(f"Erro linha {linha}: {erro}")
            else:
                sucessos += 1
                job["logs"].append(f"Linha {linha}: {resposta}")
            job["atual"] = linha
            job["sucessos"] = sucessos
            job["erros"] = erros

        time.sleep(intervalo)
        linha += 1

    saldo = None
    if usuario and senha:
        try:
            saldo = _consultar_saldo_sms(usuario, senha)
        except requests.exceptions.RequestException:
            pass

    with COMANDO_JOBS_LOCK:
        COMANDO_JOBS[job_id]["status"] = "concluido"
        COMANDO_JOBS[job_id]["saldo"] = saldo


@app.route("/api/comando/enviar-massa", methods=["POST"])
def comando_enviar_massa():
    auth = session.get("sms_auth")
    if not auth:
        return jsonify(ok=False, error="Autentique-se na SMS Market primeiro."), 401
    usuario, senha = session.get("sms_usuario"), session.get("sms_senha")

    body = request.get_json(force=True) or {}
    file_id = body.get("file_id")
    intervalo = _para_int(body.get("intervalo"))

    with UPLOADS_MASSA_LOCK:
        conteudo = UPLOADS_MASSA.pop(file_id, None)
    if conteudo is None:
        return jsonify(ok=False, error="Arquivo não encontrado. Envie novamente."), 400

    wb = openpyxl.load_workbook(io.BytesIO(conteudo))
    sheet = wb.active
    coluna1 = [str(cell.value) for cell in sheet["A"]]
    coluna2 = [str(cell.value) for cell in sheet["B"]]
    coluna3 = [str(cell.value) for cell in sheet["C"]]
    total = max(len(coluna1), 1)

    job_id = uuid.uuid4().hex
    with COMANDO_JOBS_LOCK:
        COMANDO_JOBS[job_id] = {
            "status": "rodando", "atual": 0, "total": total,
            "sucessos": 0, "erros": 0, "logs": [], "saldo": None,
        }

    threading.Thread(
        target=_executar_envio_massa,
        args=(job_id, coluna1, coluna2, coluna3, intervalo, auth, usuario, senha),
        daemon=True,
    ).start()

    return jsonify(ok=True, job_id=job_id, total=total)


@app.route("/api/comando/enviar-massa/status/<job_id>")
def comando_enviar_massa_status(job_id):
    with COMANDO_JOBS_LOCK:
        job = COMANDO_JOBS.get(job_id)
        if not job:
            return jsonify(ok=False, error="Job não encontrado."), 404
        resultado = dict(job)
    return jsonify(ok=True, **resultado)


# --- CONVERSOR KML -> SSX (Áreas/Rotas) ---
# Motor de conversão em conversorkml.py (portado do projeto standalone "Conversor
# de Áreas e Rotas"), aqui só a camada web em cima das mesmas funções puras.
CONVERSOES = {}
CONVERSOES_LOCK = threading.Lock()
CONVERSOR_TAMANHO_PARTE = 1000  # abaixo do limite de importação do SSX, pra manter os arquivos leves

# Avisos automáticos que aparecem em praticamente todo registro quando se força a
# conversão Rota -> Área (cada anel é fechado automaticamente). Sem compactar,
# um KML com milhares de Placemarks devolveria milhares de linhas "problemáticas"
# pro navegador renderizar (mesmo travamento que a tela de migração já evita).
CONVERSOR_AVISOS_COMPACTAVEIS = (
    "GeoIntegrationCode truncado",
    "Anel de área fechado automaticamente",
    "Coordenadas excedem",
)


@app.route("/api/conversor/converter", methods=["POST"])
def conversor_converter():
    if "arquivo" not in request.files:
        return jsonify(ok=False, error="Nenhum arquivo enviado."), 400
    arquivo = request.files["arquivo"]

    tipo = request.form.get("tipo", "areas")
    categoria = request.form.get("categoria", "").strip() or None
    grupo = request.form.get("grupo", "").strip() or None
    tolerancia_texto = request.form.get("tolerancia", "").strip()
    tolerancia = None
    if tolerancia_texto:
        try:
            tolerancia = int(tolerancia_texto)
        except ValueError:
            return jsonify(ok=False, error=f"Tolerância deve ser um número inteiro: '{tolerancia_texto}'"), 400

    cor_texto = request.form.get("cor", "").strip()
    cor = None
    if cor_texto:
        try:
            cor = int(cor_texto)
        except ValueError:
            return jsonify(ok=False, error=f"Cor inválida: '{cor_texto}'"), 400
        if cor not in range(1, 14):
            return jsonify(ok=False, error="Cor deve ser um código de 1 a 13 (ver tabela do manual)."), 400

    try:
        kml_text = arquivo.read().decode("utf-8")
    except UnicodeDecodeError:
        return jsonify(ok=False, error="Arquivo KML não está em UTF-8."), 400

    config = conversorkml.Config(
        categoria=categoria,
        grupo=grupo,
        tolerancia=tolerancia,
        cor=cor,
        forcar_poligono=(tipo == "areas"),
    )

    try:
        registros = conversorkml.processar(kml_text, config)
    except ValueError as e:
        return jsonify(ok=False, error=str(e)), 400

    nome_base = os.path.splitext(arquivo.filename or "conversao")[0]
    conv_id = uuid.uuid4().hex
    with CONVERSOES_LOCK:
        CONVERSOES[conv_id] = {"registros": registros, "nome_base": nome_base, "sufixo": tipo}

    n_erro = sum(1 for r in registros if r["erros"])
    n_ok = len(registros) - n_erro
    n_partes = -(-n_ok // CONVERSOR_TAMANHO_PARTE) if n_ok > CONVERSOR_TAMANHO_PARTE else 1

    avisos_compactados = {"geo_truncado": 0, "anel_fechado": 0, "coordenadas_longas": 0}
    problematicos = []
    for r in registros:
        if not r["erros"] and not r["avisos"]:
            continue
        so_compactaveis = not r["erros"] and all(
            any(chave in av for chave in CONVERSOR_AVISOS_COMPACTAVEIS) for av in r["avisos"]
        )
        if so_compactaveis:
            for av in r["avisos"]:
                if "GeoIntegrationCode truncado" in av:
                    avisos_compactados["geo_truncado"] += 1
                elif "Anel de área fechado automaticamente" in av:
                    avisos_compactados["anel_fechado"] += 1
                elif "Coordenadas excedem" in av:
                    avisos_compactados["coordenadas_longas"] += 1
            continue
        problematicos.append({
            "indice": r["indice"],
            "nome": r["nome"],
            "tipo_original": r["tipo_original"],
            "convertido": r["convertido"],
            "codigo": r["dados"].get("GeoIntegrationCode", ""),
            "erros": r["erros"],
            "avisos": r["avisos"],
        })

    # Teto de linhas detalhadas na tela: com KMLs de milhares de Placemarks e um
    # aviso genérico (ex.: descrição longa) presente em quase todo registro, a
    # lista completa ainda poderia estourar e travar o navegador ao renderizar.
    LIMITE_PROBLEMATICOS = 500
    problematicos_ocultos = max(0, len(problematicos) - LIMITE_PROBLEMATICOS)
    problematicos = problematicos[:LIMITE_PROBLEMATICOS]

    return jsonify(
        ok=True,
        conv_id=conv_id,
        total=len(registros),
        n_ok=n_ok,
        n_erro=n_erro,
        n_partes=n_partes,
        tamanho_parte=CONVERSOR_TAMANHO_PARTE,
        max_linhas_importacao=conversorkml.MAX_LINHAS_IMPORTACAO,
        avisos_compactados=avisos_compactados,
        problematicos=problematicos,
        problematicos_ocultos=problematicos_ocultos,
    )


@app.route("/api/conversor/download/<conv_id>/<formato>/<int:parte>")
def conversor_download(conv_id, formato, parte):
    with CONVERSOES_LOCK:
        dados = CONVERSOES.get(conv_id)
    if not dados:
        return jsonify(ok=False, error="Conversão não encontrada. Converta novamente."), 404
    if formato not in ("kml", "csv"):
        return jsonify(ok=False, error="Formato inválido."), 400

    validos = [r for r in dados["registros"] if not r["erros"]]
    inicio = (parte - 1) * CONVERSOR_TAMANHO_PARTE
    fatia = validos[inicio:inicio + CONVERSOR_TAMANHO_PARTE]
    if not fatia:
        return jsonify(ok=False, error="Parte não encontrada."), 404

    sufixo = dados["sufixo"]
    n_partes = -(-len(validos) // CONVERSOR_TAMANHO_PARTE) if len(validos) > CONVERSOR_TAMANHO_PARTE else 1
    sufixo_parte = f"_parte{parte}" if n_partes > 1 else ""

    if formato == "kml":
        conteudo = conversorkml.gerar_kml(fatia).encode("utf-8")
        mimetype = "application/vnd.google-earth.kml+xml"
        nome_arquivo = f"{dados['nome_base']}_SSX_{sufixo}{sufixo_parte}.kml"
    else:
        conteudo = conversorkml.gerar_csv(fatia).encode("ascii", errors="ignore")
        mimetype = "text/csv"
        nome_arquivo = f"{dados['nome_base']}_SSX_{sufixo}{sufixo_parte}.csv"

    return Response(
        conteudo,
        mimetype=mimetype,
        headers={"Content-Disposition": f'attachment; filename="{nome_arquivo}"'},
    )


if __name__ == "__main__":
    app.run(host="127.0.0.1", port=5000, debug=False)
